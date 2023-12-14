import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# 엑셀 파일 생성
workbook = Workbook()

# 활성 시트 선택
sheet = workbook.active
sheet.title = 'Product Sales Data'

# 헤더 생성
headers = ['제품ID', '제품명', '수량', '가격']
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header).font = Font(bold=True)

# 데이터 생성 및 입력
for row_num in range(2, 102):
    product_id = f'P{row_num - 1:03d}'
    product_name = f'제품{row_num - 1}'
    quantity = 10 + row_num
    price = 1000 + row_num * 10

    sheet.cell(row=row_num, column=1, value=product_id)
    sheet.cell(row=row_num, column=2, value=product_name)
    sheet.cell(row=row_num, column=3, value=quantity)
    sheet.cell(row=row_num, column=4, value=price)

# 스크립트 실행 디렉토리에 엑셀 파일 저장
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'products.xlsx')
workbook.save(file_path)

print(f'데이터가 {file_path}에 성공적으로 저장되었습니다.')
